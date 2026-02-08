from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-this')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24 * 30  # 30 days

security = HTTPBearer()

# Create the main app without a prefix
app = FastAPI()

# CORS: middleware that adds headers to EVERY response (no dependency on Starlette CORS)
class AddCORSHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        if request.method == "OPTIONS":
            from starlette.responses import Response
            return Response(
                status_code=200,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, PATCH, OPTIONS",
                    "Access-Control-Allow-Headers": "*",
                    "Access-Control-Max-Age": "86400",
                },
            )
        response = await call_next(request)
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, PATCH, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "*"
        return response


app.add_middleware(AddCORSHeadersMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class UserRegister(BaseModel):
    name: str
    email: EmailStr
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: EmailStr
    created_at: datetime

class AuthResponse(BaseModel):
    token: str
    user: User

class ExpenseCreate(BaseModel):
    amount: float
    category: str
    description: str
    date: str  # YYYY-MM-DD format

class ExpenseUpdate(BaseModel):
    amount: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    amount: float
    category: str
    description: str
    date: str
    created_at: datetime

class MonthlySummary(BaseModel):
    total_expenses: float
    total_count: int
    category_breakdown: List[Dict[str, Any]]
    daily_expenses: List[Dict[str, Any]]
    top_categories: List[Dict[str, Any]]

# Helper functions
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str) -> str:
    payload = {
        'user_id': user_id,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> str:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        return user_id
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Auth routes (with and without trailing slash to avoid 405 on redirect)
@api_router.get("/auth/register")
@api_router.get("/auth/login")
async def auth_get_hint():
    """Opening this URL in a browser sends GET; use POST from the app with JSON body."""
    return {"detail": "Use POST with JSON body (name, email, password for register; email, password for login)."}


@api_router.post("/auth/register", response_model=AuthResponse)
@api_router.post("/auth/register/", response_model=AuthResponse)
async def register(user_data: UserRegister):
    # Check if user exists
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "name": user_data.name,
        "email": user_data.email,
        "password_hash": hash_password(user_data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    # Create token
    token = create_token(user_id)
    
    user = User(
        id=user_id,
        name=user_data.name,
        email=user_data.email,
        created_at=datetime.fromisoformat(user_doc['created_at'])
    )
    
    return AuthResponse(token=token, user=user)

@api_router.post("/auth/login", response_model=AuthResponse)
@api_router.post("/auth/login/", response_model=AuthResponse)
async def login(user_data: UserLogin):
    # Find user
    user_doc = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if not user_doc or not verify_password(user_data.password, user_doc['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Create token
    token = create_token(user_doc['id'])
    
    user = User(
        id=user_doc['id'],
        name=user_doc['name'],
        email=user_doc['email'],
        created_at=datetime.fromisoformat(user_doc['created_at'])
    )
    
    return AuthResponse(token=token, user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(user_id: str = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    return User(
        id=user_doc['id'],
        name=user_doc['name'],
        email=user_doc['email'],
        created_at=datetime.fromisoformat(user_doc['created_at'])
    )

# Expense routes
@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense_data: ExpenseCreate, user_id: str = Depends(get_current_user)):
    expense_id = str(uuid.uuid4())
    expense_doc = {
        "id": expense_id,
        "user_id": user_id,
        "amount": expense_data.amount,
        "category": expense_data.category,
        "description": expense_data.description,
        "date": expense_data.date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.expenses.insert_one(expense_doc)
    
    # Convert created_at string back to datetime for the response model
    expense_doc['created_at'] = datetime.fromisoformat(expense_doc['created_at'])
    return Expense(**expense_doc)

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(
    user_id: str = Depends(get_current_user),
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {"user_id": user_id}
    
    if category:
        query["category"] = category
    
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    for expense in expenses:
        if isinstance(expense['created_at'], str):
            expense['created_at'] = datetime.fromisoformat(expense['created_at'])
    
    return expenses

@api_router.get("/expenses/{expense_id}", response_model=Expense)
async def get_expense(expense_id: str, user_id: str = Depends(get_current_user)):
    expense = await db.expenses.find_one({"id": expense_id, "user_id": user_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if isinstance(expense['created_at'], str):
        expense['created_at'] = datetime.fromisoformat(expense['created_at'])
    
    return Expense(**expense)

@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(
    expense_id: str,
    expense_data: ExpenseUpdate,
    user_id: str = Depends(get_current_user)
):
    expense = await db.expenses.find_one({"id": expense_id, "user_id": user_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    update_data = expense_data.model_dump(exclude_unset=True)
    if update_data:
        await db.expenses.update_one({"id": expense_id}, {"$set": update_data})
        expense.update(update_data)
    
    if isinstance(expense['created_at'], str):
        expense['created_at'] = datetime.fromisoformat(expense['created_at'])
    
    return Expense(**expense)

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, user_id: str = Depends(get_current_user)):
    result = await db.expenses.delete_one({"id": expense_id, "user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted"}

@api_router.get("/expenses/summary/monthly", response_model=MonthlySummary)
async def get_monthly_summary(
    month: str,  # Format: YYYY-MM
    user_id: str = Depends(get_current_user)
):
    # Calculate date range
    year, month_num = map(int, month.split('-'))
    start_date = f"{year}-{month_num:02d}-01"
    
    # Calculate last day of month
    if month_num == 12:
        next_month = f"{year + 1}-01-01"
    else:
        next_month = f"{year}-{month_num + 1:02d}-01"
    
    # Query expenses for the month
    expenses = await db.expenses.find({
        "user_id": user_id,
        "date": {"$gte": start_date, "$lt": next_month}
    }, {"_id": 0}).to_list(10000)
    
    # Calculate totals
    total_expenses = sum(exp['amount'] for exp in expenses)
    total_count = len(expenses)
    
    # Category breakdown
    category_totals = {}
    for exp in expenses:
        cat = exp['category']
        category_totals[cat] = category_totals.get(cat, 0) + exp['amount']
    
    category_breakdown = [
        {"category": cat, "amount": amt, "percentage": round((amt / total_expenses * 100) if total_expenses > 0 else 0, 2)}
        for cat, amt in category_totals.items()
    ]
    category_breakdown.sort(key=lambda x: x['amount'], reverse=True)
    
    # Daily expenses
    daily_totals = {}
    for exp in expenses:
        date = exp['date']
        daily_totals[date] = daily_totals.get(date, 0) + exp['amount']
    
    daily_expenses = [
        {"date": date, "amount": amt}
        for date, amt in sorted(daily_totals.items())
    ]
    
    # Top categories (top 5)
    top_categories = category_breakdown[:5]
    
    return MonthlySummary(
        total_expenses=total_expenses,
        total_count=total_count,
        category_breakdown=category_breakdown,
        daily_expenses=daily_expenses,
        top_categories=top_categories
    )

@api_router.get("/expenses/export/pdf")
async def export_pdf(
    month: str,  # Format: YYYY-MM
    user_id: str = Depends(get_current_user)
):
    # Get expenses
    year, month_num = map(int, month.split('-'))
    start_date = f"{year}-{month_num:02d}-01"
    if month_num == 12:
        next_month = f"{year + 1}-01-01"
    else:
        next_month = f"{year}-{month_num + 1:02d}-01"
    
    expenses = await db.expenses.find({
        "user_id": user_id,
        "date": {"$gte": start_date, "$lt": next_month}
    }, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"<b>Expense Report - {month}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Summary
    total = sum(exp['amount'] for exp in expenses)
    summary = Paragraph(f"<b>Total Expenses: ₹{total:,.2f}</b>", styles['Heading2'])
    elements.append(summary)
    elements.append(Spacer(1, 20))
    
    # Table
    table_data = [['Date', 'Category', 'Description', 'Amount (₹)']]
    for exp in expenses:
        table_data.append([
            exp['date'],
            exp['category'],
            exp['description'][:30],
            f"₹{exp['amount']:,.2f}"
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=expenses_{month}.pdf"}
    )

@api_router.get("/expenses/export/excel")
async def export_excel(
    month: str,  # Format: YYYY-MM
    user_id: str = Depends(get_current_user)
):
    # Get expenses
    year, month_num = map(int, month.split('-'))
    start_date = f"{year}-{month_num:02d}-01"
    if month_num == 12:
        next_month = f"{year + 1}-01-01"
    else:
        next_month = f"{year}-{month_num + 1:02d}-01"
    
    expenses = await db.expenses.find({
        "user_id": user_id,
        "date": {"$gte": start_date, "$lt": next_month}
    }, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Headers
    headers = ['Date', 'Category', 'Description', 'Amount (₹)']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    
    # Data
    for exp in expenses:
        ws.append([exp['date'], exp['category'], exp['description'], exp['amount']])
    
    # Summary
    ws.append([])
    total = sum(exp['amount'] for exp in expenses)
    ws.append(['Total', '', '', total])
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=expenses_{month}.xlsx"}
    )

# Root and health (avoid 404 when someone opens backend URL in browser)
@app.get("/")
async def root():
    return {"message": "VividExpense API", "docs": "/docs", "api": "/api"}


@api_router.get("/health")
async def health():
    return {"status": "ok"}


# Include the router in the main app
app.include_router(api_router)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()